import time
import pandas as pd
import openpyxl
from engine import mapeamento_loader, transformador, exportador
from engine.logger import get_logger
from models.schemas import ResultadoProcessamento

_log = get_logger("processador")


def _ler_excel_robusto(caminho: str) -> pd.DataFrame:
    """
    Lê um arquivo Excel tratando células mescladas e cabeçalho fora da linha 1.

    Estratégia:
    1. Usa openpyxl para desunificar células mescladas (preenche o valor
       da célula âncora em todas as células da região).
    2. Varre as primeiras 15 linhas para achar a linha de cabeçalho:
       a linha com mais células preenchidas e sem valores tipo "Unnamed".
    3. Lê o DataFrame a partir dessa linha.
    """
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active

    # ── Desunifica células mescladas ─────────────────────────────────────────
    for merged in list(ws.merged_cells.ranges):
        min_row, min_col = merged.min_row, merged.min_col
        valor_ancora = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merged))
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=merged.max_row,
            min_col=min_col,
            max_col=merged.max_col,
        ):
            for cell in row:
                cell.value = valor_ancora

    # ── Converte para lista de listas ────────────────────────────────────────
    dados = [[cell.value for cell in row] for row in ws.iter_rows()]
    wb.close()

    if not dados:
        return pd.DataFrame()

    # ── Detecta linha de cabeçalho (máx 15 linhas de prospecção) ─────────────
    n_cols = max(len(r) for r in dados)
    melhor_linha = 0
    melhor_score = -1

    for i, linha in enumerate(dados[:15]):
        # score = n° de células com texto não nulo e não "Unnamed"
        score = sum(
            1
            for v in linha
            if v is not None
            and str(v).strip() != ""
            and not str(v).startswith("Unnamed")
        )
        if score > melhor_score:
            melhor_score = score
            melhor_linha = i

    cabecalho = [str(v).strip() if v is not None else "" for v in dados[melhor_linha]]

    # garante nomes únicos para colunas sem nome
    seen: dict = {}
    cabecalho_unico = []
    for nome in cabecalho:
        if nome == "" or nome == "None":
            nome = "_COL_VAZIA_"
        if nome in seen:
            seen[nome] += 1
            nome = f"{nome}_{seen[nome]}"
        else:
            seen[nome] = 0
        cabecalho_unico.append(nome)

    linhas_dados = dados[melhor_linha + 1 :]
    df = pd.DataFrame(linhas_dados, columns=cabecalho_unico, dtype=object)

    # remove colunas completamente vazias geradas por células mescladas de layout
    df = df.loc[:, ~df.columns.str.startswith("_COL_VAZIA_")]
    # remove linhas 100% vazias
    df = df.dropna(how="all").reset_index(drop=True)

    # converte tudo para string (compatível com o resto do pipeline)
    return df.astype(str).replace("None", "").replace("nan", "")


def preview_base(caminho_arquivo: str, cod_varejista: int, n_linhas: int = 10) -> dict:
    """
    Executa as transformações nas primeiras `n_linhas` da base sem exportar.
    Retorna {"ok": bool, "colunas": list, "linhas": list[list], "erro": str|None}
    """
    try:
        mapeamento = mapeamento_loader.carregar(cod_varejista)
        if not mapeamento:
            return {
                "ok": False,
                "colunas": [],
                "linhas": [],
                "erro": "Mapeamento não configurado.",
            }

        df = _ler_excel_robusto(caminho_arquivo)
        df.columns = df.columns.str.strip()
        df = df.head(n_linhas)

        ignorar = [c for c in mapeamento.get("ignorar", []) if c in df.columns]
        if ignorar:
            df.drop(columns=ignorar, inplace=True)

        if mapeamento.get("separar"):
            df = transformador.separar_mes_ano(df, mapeamento["separar"])

        if mapeamento.get("cruzar_varejista"):
            df, _ = transformador.cruzar_varejista(df, mapeamento["cruzar_varejista"])

        if mapeamento.get("cruzar_loja"):
            df, _, _ = transformador.cruzar_loja(
                df, mapeamento["cruzar_loja"], cod_varejista
            )

        if mapeamento.get("cruzar_ean"):
            df = transformador.cruzar_ean(df, mapeamento["cruzar_ean"])

        if mapeamento.get("renomear"):
            df = transformador.renomear_colunas(df, mapeamento["renomear"])

        colunas_calc = set(mapeamento.get("calcular", {}).keys())
        df = transformador.converter_numericos(df, colunas_calc)

        if mapeamento.get("calcular"):
            df = transformador.calcular_colunas(df, mapeamento["calcular"])

        if mapeamento.get("novas"):
            df = transformador.adicionar_colunas_novas(
                df, mapeamento["novas"], mapeamento.get("renomear", {})
            )

        # remove colunas temporárias internas
        for tmp in ("_LOJA_OK_", "_COD_VAR_"):
            if tmp in df.columns:
                df.drop(columns=[tmp], inplace=True)

        colunas = list(df.columns)
        linhas = df.fillna("").astype(str).values.tolist()
        return {"ok": True, "colunas": colunas, "linhas": linhas, "erro": None}

    except Exception as e:
        _log.error("Erro no dry-run: %s", e, exc_info=True)
        return {"ok": False, "colunas": [], "linhas": [], "erro": str(e)}


def processar_base(
    caminho_arquivo: str,
    cod_varejista: int,
    nome_varejista: str,
    on_status=None,
) -> dict:
    """
    Orquestra o processamento completo de uma base Excel.

    Retorna dict com:
    - ok, erro
    - arquivo_saida
    - total_linhas
    - lojas_unicas      ← total de lojas únicas na base
    - lojas_ok          ← lojas identificadas no banco
    - lojas_novas       ← lojas não encontradas
    - total_valor       ← soma da coluna VALOR
    - total_quantidade  ← soma da coluna QUANTIDADE
    - setores           ← lista de setores únicos de SETOR_PRODUTO
    - pendencias
    """

    # Pesos cumulativos de progresso (0.0 – 1.0) por etapa
    _PESOS = {
        "Carregando mapeamento...": 0.05,
        "Lendo arquivo Excel...": 0.15,
        "Removendo colunas ignoradas...": 0.17,
        "Separando data em MÊS e ANO...": 0.20,
        "Identificando lojas...": 0.58,
        "Cruzando varejistas...": 0.64,
        "Cruzando EANs...": 0.70,
        "Renomeando colunas...": 0.72,
        "Convertendo valores numéricos...": 0.78,
        "Calculando colunas...": 0.82,
        "Adicionando novas colunas...": 0.85,
        "Sinalizando pendências...": 0.88,
        "Exportando arquivo Excel...": 1.00,
    }

    def _status(msg: str):
        if on_status:
            try:
                on_status(msg, _PESOS.get(msg))
            except Exception:
                pass

    timings = {}
    inicio_total = time.perf_counter()

    # ── 1. Mapeamento ─────────────────────────────────────────────────────────
    _status("Carregando mapeamento...")
    inicio = time.perf_counter()
    mapeamento = mapeamento_loader.carregar(cod_varejista)
    timings["load_mapping"] = time.perf_counter() - inicio

    if not mapeamento:
        _log.warning(
            "Mapeamento não encontrado para varejista %s (cod=%s)",
            nome_varejista,
            cod_varejista,
        )
        return {
            "ok": False,
            "erro": f"Nenhum mapeamento configurado para '{nome_varejista}'. "
            f"Configure em Configurar Mapeamento.",
        }

    # ── 2. Leitura ────────────────────────────────────────────────────────────
    try:
        _status("Lendo arquivo Excel...")
        inicio = time.perf_counter()
        df = _ler_excel_robusto(caminho_arquivo)
        df.columns = df.columns.str.strip()
        timings["read_excel"] = time.perf_counter() - inicio
    except Exception as e:
        _log.error("Erro ao ler Excel '%s': %s", caminho_arquivo, e, exc_info=True)
        return {"ok": False, "erro": f"Erro ao ler o arquivo: {e}"}

    total_linhas = len(df)

    # ── 3. Colunas a ignorar (remove antes de processar) ──────────────────────
    _status("Removendo colunas ignoradas...")
    inicio = time.perf_counter()
    colunas_ignorar = [
        col for col in mapeamento.get("ignorar", []) if col in df.columns
    ]
    if colunas_ignorar:
        df.drop(columns=colunas_ignorar, inplace=True)
    timings["drop_ignored"] = time.perf_counter() - inicio

    # ── 4. Transformações ─────────────────────────────────────────────────────

    if mapeamento.get("separar"):
        _status("Separando data em MÊS e ANO...")
        inicio = time.perf_counter()
        df = transformador.separar_mes_ano(df, mapeamento["separar"])
        timings["separar_mes_ano"] = time.perf_counter() - inicio

    pendencias = []
    saida_id = "LOJA"
    saida_nome = "BANCO"
    varejistas_novos = []
    # cruzar_varejista vem ANTES de cruzar_loja para popular _COD_VAR_ por linha
    if mapeamento.get("cruzar_varejista"):
        _status("Cruzando varejistas...")
        inicio = time.perf_counter()
        df, varejistas_novos = transformador.cruzar_varejista(
            df, mapeamento["cruzar_varejista"]
        )
        timings["cruzar_varejista"] = time.perf_counter() - inicio

    if mapeamento.get("cruzar_loja"):
        _status("Identificando lojas...")
        inicio = time.perf_counter()
        df, pendencias, saida_id = transformador.cruzar_loja(
            df, mapeamento["cruzar_loja"], cod_varejista
        )
        timings["cruzar_loja"] = time.perf_counter() - inicio
        saida_nome = mapeamento["cruzar_loja"].get("saida_nome", "BANCO")
        # remove coluna temporária de varejista por linha
        if "_COD_VAR_" in df.columns:
            df.drop(columns=["_COD_VAR_"], inplace=True)

    # garante que _COD_VAR_ é removida mesmo sem cruzar_loja
    if "_COD_VAR_" in df.columns:
        df.drop(columns=["_COD_VAR_"], inplace=True)

    if mapeamento.get("cruzar_ean"):
        _status("Cruzando EANs...")
        inicio = time.perf_counter()
        df = transformador.cruzar_ean(df, mapeamento["cruzar_ean"])
        timings["cruzar_ean"] = time.perf_counter() - inicio

    if mapeamento.get("renomear"):
        _status("Renomeando colunas...")
        inicio = time.perf_counter()
        df = transformador.renomear_colunas(df, mapeamento["renomear"])
        timings["renomear"] = time.perf_counter() - inicio

    colunas_calculadas = set(mapeamento.get("calcular", {}).keys())
    _status("Convertendo valores numéricos...")
    df = transformador.converter_numericos(df, {saida_id} | colunas_calculadas)

    if mapeamento.get("calcular"):
        _status("Calculando colunas...")
        inicio = time.perf_counter()
        df = transformador.calcular_colunas(df, mapeamento["calcular"])
        timings["calcular"] = time.perf_counter() - inicio

    if mapeamento.get("novas"):
        _status("Adicionando novas colunas...")
        inicio = time.perf_counter()
        df = transformador.adicionar_colunas_novas(
            df, mapeamento["novas"], mapeamento.get("renomear", {})
        )
        timings["adicionar_novas"] = time.perf_counter() - inicio

    _status("Sinalizando pendências...")
    inicio = time.perf_counter()
    df = transformador.sinalizar_pendencias(df, pendencias, saida_nome)
    timings["sinalizar_pendencias"] = time.perf_counter() - inicio

    # ── 5. Estatísticas ───────────────────────────────────────────────────────

    # total valor — busca qualquer coluna com "valor" no nome
    total_valor = 0.0
    col_valor = next((c for c in df.columns if "valor" in c.lower()), None)
    if col_valor:
        _v = pd.to_numeric(df[col_valor], errors="coerce").sum()
        total_valor = round(float(_v), 2) if not pd.isna(_v) else 0.0

    # total quantidade — busca qualquer coluna com "quant" ou "qtd" no nome
    total_quantidade = 0.0
    col_qtd = next(
        (c for c in df.columns if "quant" in c.lower() or "qtd" in c.lower()), None
    )
    if col_qtd:
        _q = pd.to_numeric(df[col_qtd], errors="coerce").sum()
        total_quantidade = round(float(_q), 2) if not pd.isna(_q) else 0.0

    # setores únicos
    setores = []
    if "SETOR_PRODUTO" in df.columns:
        setores = [
            s
            for s in df["SETOR_PRODUTO"].dropna().unique().tolist()
            if s and s != "NÃO IDENTIFICADO"
        ]
        setores = sorted(setores)

    # lojas únicas identificadas
    lojas_unicas = 0
    lojas_ok = 0
    if saida_id in df.columns:
        lojas_unicas = df[saida_id].nunique()

    if saida_nome in df.columns and saida_id in df.columns:
        lojas_ok = int(
            df.loc[df[saida_nome] != "NÃO ENCONTRADO", saida_id]
            .dropna()
            .astype(str)
            .nunique()
        )

    lojas_novas = len(pendencias)

    # mês + ano referência para nome do arquivo de saída
    mes_ref = ""
    _mes_val = ""
    _ano_val = ""
    for col_mes in ("MÊS", "MES"):
        if col_mes in df.columns:
            vals = [
                v
                for v in df[col_mes].dropna().unique()
                if str(v).strip() not in ("", "nan")
            ]
            if vals:
                _mes_val = str(vals[0])
                break
    if "ANO" in df.columns:
        vals = [
            v for v in df["ANO"].dropna().unique() if str(v).strip() not in ("", "nan")
        ]
        if vals:
            ano_str = str(vals[0])
            # normaliza 2 dígitos → 4 dígitos
            if len(ano_str) == 2 and ano_str.isdigit():
                ano_str = "20" + ano_str
            _ano_val = ano_str

    if _mes_val and _ano_val:
        mes_ref = f"{_mes_val}_{_ano_val}"
    elif _mes_val:
        mes_ref = _mes_val
    elif _ano_val:
        mes_ref = _ano_val

    # coluna de varejista cruzado (para permitir split por varejista no download)
    coluna_varejista_saida = ""
    if mapeamento.get("cruzar_varejista"):
        coluna_varejista_saida = mapeamento["cruzar_varejista"].get("saida", "")

    # ── 6. Exportar ───────────────────────────────────────────────────────────
    _status("Exportando arquivo Excel...")
    try:
        inicio = time.perf_counter()
        arquivo_saida = exportador.salvar_excel(df, pendencias, nome_varejista)
        timings["exportar"] = time.perf_counter() - inicio
    except Exception as e:
        _log.error(
            "Erro ao exportar Excel para varejista %s: %s",
            nome_varejista,
            e,
            exc_info=True,
        )
        return {"ok": False, "erro": f"Erro ao salvar: {e}"}

    timings["total"] = time.perf_counter() - inicio_total

    # log de timings para diagnóstico de gargalos em produção
    _log.info(
        "Processamento '%s' concluído em %.2fs | linhas=%d lojas_ok=%d pendencias=%d | etapas: %s",
        nome_varejista,
        timings["total"],
        total_linhas,
        lojas_ok,
        lojas_novas,
        " | ".join(f"{k}={v:.3f}s" for k, v in timings.items() if k != "total"),
    )

    return ResultadoProcessamento(
        ok=True,
        arquivo_saida=arquivo_saida,
        total_linhas=total_linhas,
        lojas_unicas=int(lojas_unicas),
        lojas_ok=lojas_ok,
        lojas_novas=lojas_novas,
        total_valor=total_valor,
        total_quantidade=total_quantidade,
        setores=setores,
        pendencias=pendencias,
        varejistas_novos=varejistas_novos,
        mes_ref=mes_ref,
        coluna_varejista_saida=coluna_varejista_saida,
        erro=None,
        timings=timings,
    ).model_dump()
